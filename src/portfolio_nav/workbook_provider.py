"""Load monthly NAV series from Ahil-filled xlsx workbooks (interim until nav_engine.py)."""

from __future__ import annotations

import logging
from functools import lru_cache
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook

from src.config_paths import BASE_DIR
from src.portfolio_nav.stats import build_nav_points, monthly_return_pcts
from src.portfolio_nav.attribution import attribution_for_book, load_nav_config
from src.portfolio_nav.types import MonthlyReturn, NavHistoryBundle, NavPoint

logger = logging.getLogger(__name__)

_MONTHLY_NAV_SHEET = "Monthly_NAV"
_HEADER_ROW = 3
_DATA_START_ROW = 4
_COL_MONTH = 1
_COL_OPEN = 2
_COL_CLOSE = 3
_COL_BENCH = 11


def _load_nav_config() -> dict[str, Any]:
    return load_nav_config()


def _workbook_path(key: str) -> Path:
    cfg = _load_nav_config()
    rel = ((cfg.get("workbooks") or {}).get(key) or {}).get("path")
    if not rel:
        raise FileNotFoundError(f"Workbook path for '{key}' not configured in portfolio_nav.yaml")
    path = BASE_DIR / rel
    if not path.is_file():
        raise FileNotFoundError(f"Ahil NAV workbook not found: {path}")
    return path


def _parse_monthly_sheet(path: Path) -> tuple[list[str], list[float], list[float | None]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[_MONTHLY_NAV_SHEET]
        labels: list[str] = []
        closes: list[float] = []
        bench: list[float | None] = []
        for row in ws.iter_rows(min_row=_DATA_START_ROW, values_only=True):
            month = row[_COL_MONTH] if len(row) > _COL_MONTH else None
            close = row[_COL_CLOSE] if len(row) > _COL_CLOSE else None
            bench_val = row[_COL_BENCH] if len(row) > _COL_BENCH else None
            if not month or close is None:
                continue
            try:
                close_f = float(close)
            except (TypeError, ValueError):
                continue
            labels.append(str(month).strip())
            closes.append(close_f)
            bench.append(float(bench_val) if bench_val is not None else None)
        return labels, closes, bench
    finally:
        wb.close()


def _points_from_dicts(rows: list[dict[str, float | str]]) -> list[NavPoint]:
    return [
        NavPoint(
            date=str(r["date"]),
            value=float(r["value"]),
            drawdown_pct=float(r["drawdown_pct"]),
            high_water_mark=float(r["high_water_mark"]),
        )
        for r in rows
    ]


def _attribution_for_book(book: str) -> list:
    return attribution_for_book(book)


@lru_cache(maxsize=4)
def _cached_workbook_pair(mtm_mtime: float, closed_mtime: float) -> tuple[list[str], list[float], list[float], list[float | None]]:
    del mtm_mtime, closed_mtime  # cache key only
    mtm_path = _workbook_path("mtm")
    closed_path = _workbook_path("closed")
    labels_m, closes_m, bench = _parse_monthly_sheet(mtm_path)
    labels_c, closes_c, _ = _parse_monthly_sheet(closed_path)
    if labels_m != labels_c:
        logger.warning("MTM and closed workbook month labels differ; using MTM label set")
    return labels_m, closes_m, closes_c, bench


def load_workbook_history(book: str) -> NavHistoryBundle:
    """Parse Ahil xlsx workbooks into a NavHistoryBundle for any MODEL valuation book."""
    cfg = _load_nav_config()
    mtm_path = _workbook_path("mtm")
    closed_path = _workbook_path("closed")
    labels, closes_m, closes_c, bench_decimals = _cached_workbook_pair(
        mtm_path.stat().st_mtime,
        closed_path.stat().st_mtime,
    )
    if not labels:
        raise FileNotFoundError("No monthly NAV rows found in Ahil workbook")

    inception = float(cfg.get("research_notional_usd") or 10_000_000)
    monthly_pcts = monthly_return_pcts(closes_m)
    monthly_returns = [
        MonthlyReturn(month=_month_label_to_yyyy_mm(lbl), return_pct=round(pct, 2))
        for lbl, pct in zip(labels[1:], monthly_pcts)
    ]

    bench_closes: list[float] = []
    if bench_decimals and labels:
        bench_start = inception
        bench_closes.append(bench_start)
        for dec in bench_decimals[1:]:
            if dec is None:
                bench_closes.append(bench_closes[-1])
            else:
                bench_closes.append(bench_closes[-1] * (1.0 + float(dec)))

    mtm_points = _points_from_dicts(build_nav_points(labels, closes_m))
    closed_points = _points_from_dicts(build_nav_points(labels, closes_c))
    bench_points = _points_from_dicts(build_nav_points(labels, bench_closes)) if bench_closes else []

    version_m = ((cfg.get("workbooks") or {}).get("mtm") or {}).get("version_label", "MTM")
    version_c = ((cfg.get("workbooks") or {}).get("closed") or {}).get("version_label", "closed")

    note = (
        f"Monthly series from Ahil workbook ({version_m} / {version_c}). "
        f"book={book} uses proxy attribution. "
        "Daily NAV (mtm_daily[]) requires nav_engine — empty on workbook-only path."
    )

    return NavHistoryBundle(
        book=book,
        source="workbook",
        inception_nav=inception,
        mtm=mtm_points,
        closed=closed_points,
        benchmark=bench_points,
        monthly_returns=monthly_returns,
        attribution=_attribution_for_book(book),
        position_limit=int(cfg.get("position_limit_n") or 60),
        nav_history_note=note,
        metadata={
            "mtm_workbook": str(mtm_path.relative_to(BASE_DIR)),
            "closed_workbook": str(closed_path.relative_to(BASE_DIR)),
            "month_count": len(labels),
            "research_notional_usd": inception,
        },
    )


def workbook_available() -> bool:
    try:
        _workbook_path("mtm")
        _workbook_path("closed")
        return True
    except FileNotFoundError:
        return False


def _month_label_to_yyyy_mm(mmm_yy: str) -> str:
    from datetime import datetime

    try:
        dt = datetime.strptime(mmm_yy.strip(), "%b-%y")
        return f"{dt.year:04d}-{dt.month:02d}"
    except ValueError:
        return mmm_yy
